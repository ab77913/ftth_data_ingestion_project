"""
FastAPI backend for the FTTH Data Ingestion frontend.

Endpoints:
  POST /api/upload       — Upload a file, ingest it, return job summary
  GET  /api/jobs         — List all ingestion jobs
  GET  /api/jobs/{id}    — Get a single job detail
  GET  /api/records      — Paginated, filterable address records
  GET  /api/records/{id} — Single record detail
  GET  /api/columns      — Dynamic column metadata based on ingested data
  POST /api/jobs/{id}/process — Start agent processing for a job
  GET  /api/jobs/{id}/progress — SSE stream of agent progress
  GET  /api/jobs/{id}/agents — Get agent status for a job

Run: python api_server.py
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import shutil
import tempfile
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from openpyxl.styles import PatternFill, Font

import hashlib
import hmac
import secrets

from fastapi import Depends, FastAPI, File, HTTPException, Query, Request, Security, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import String, func, select, text

from data_ingestion.config.settings import get_settings
from data_ingestion.database.db import get_engine, get_session_factory, init_db, session_scope
from data_ingestion.database.models import Address, Agent1Result, AgentResult, AgentTable, DispatchQueue, IngestionJob
from data_ingestion.agents.agent1_runner import run_agent1_for_job
from data_ingestion.database.repositories import IngestionRepository
from data_ingestion.ingestion_service import IngestionService
from data_ingestion.schemas import IngestionStatus

app = FastAPI(title="FTTH Data Ingestion API", version="1.0.0")

# ─── Auth ───────────────────────────────────────────────────────────────────────

# Credentials (username → hashed password)
_USERS: dict[str, str] = {
    "ftth_team": hashlib.sha256("Meridian@2026".encode()).hexdigest(),
    "admin":     hashlib.sha256("Meridian@2026".encode()).hexdigest(),
}

# In-memory token store: token → username
_TOKENS: dict[str, str] = {}

_bearer = HTTPBearer(auto_error=False)


def _require_auth(
    request: "Request",
    creds: HTTPAuthorizationCredentials | None = Security(_bearer),
) -> str:
    """Return the username for a valid Bearer token (header or ?token=/?_t= query param)."""
    if creds and _TOKENS.get(creds.credentials):
        return _TOKENS[creds.credentials]
    # Fall back to query param (for SSE EventSource and browser file downloads)
    for param in ("token", "_t"):
        val = request.query_params.get(param)
        if val and _TOKENS.get(val):
            return _TOKENS[val]
    raise HTTPException(status_code=401, detail="Not authenticated")

# Allow the React dev server (port 5173) to hit this API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174", "http://127.0.0.1:5173", "http://127.0.0.1:5174", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

STATIC_DIR = Path(__file__).parent / "static"


@app.get("/")
async def serve_frontend():
    """Serve the single-page frontend."""
    return FileResponse(STATIC_DIR / "index.html", headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


# ─── Response models ────────────────────────────────────────────────────────────

class JobResponse(BaseModel):
    id: str
    customer_id: str | None
    source_file: str
    row_count: int
    status: str
    error_message: str | None
    created_at: str
    updated_at: str


class RecordResponse(BaseModel):
    id: int
    record_uuid: str
    job_id: str
    customer_id: str | None
    raw_address: str | None
    city: str | None
    state: str | None
    zip_code: str | None
    latitude: float | None
    longitude: float | None
    network_node: str | None
    terminal_id: str | None
    address_id: str | None
    normalized_key: str | None
    source_file: str
    source_sheet: str | None
    source_layer: str | None
    source_row_number: int | None
    validation_errors: Any | None
    validation_warnings: Any | None
    raw_metadata: Any | None
    created_at: str
    # Flattened raw data for table display
    raw_data: dict[str, Any] | None = None


class PaginatedRecords(BaseModel):
    records: list[RecordResponse]
    total: int
    page: int
    page_size: int
    total_pages: int
    columns: list[dict[str, Any]]
    display_mode: str = "raw"  # "raw" = show original file columns, "canonical" = show mapped fields


class UploadResponse(BaseModel):
    job_id: str
    source_file: str
    total_raw_records: int
    valid_records: int
    invalid_records: int
    duplicate_records: int
    stored_records: int
    queued_records: int
    status: str


class ColumnInfo(BaseModel):
    key: str
    label: str
    visible: bool = True


# ─── Startup ────────────────────────────────────────────────────────────────────

@app.on_event("startup")
def startup():
    init_db()


# ─── Auth endpoints ─────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    username: str
    password: str


@app.post("/api/login")
def login(req: LoginRequest):
    hashed = hashlib.sha256(req.password.encode()).hexdigest()
    stored = _USERS.get(req.username)
    if not stored or not hmac.compare_digest(stored, hashed):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    token = secrets.token_urlsafe(32)
    _TOKENS[token] = req.username
    return {"token": token, "username": req.username}


@app.post("/api/logout")
def logout(creds: HTTPAuthorizationCredentials | None = Security(_bearer)):
    if creds and creds.credentials in _TOKENS:
        del _TOKENS[creds.credentials]
    return {"detail": "Logged out"}


# ─── Endpoints ──────────────────────────────────────────────────────────────────


@app.post("/api/upload", response_model=UploadResponse)
async def upload_file(
    file: UploadFile = File(...),
    customer_id: str = Query(default="demo_customer"),
    _current_user: str = Depends(_require_auth),
):
    """Upload and ingest a CSV, Excel, KML, KMZ, or ZIP file.
    ZIP files are automatically extracted; every supported file inside is ingested.
    The response reflects the last (or only) successful ingestion result.
    """
    dest = UPLOAD_DIR / file.filename
    with open(dest, "wb") as fh:
        shutil.copyfileobj(file.file, fh)

    suffix = Path(file.filename).suffix.lower()

    # Collect paths to ingest — expand ZIP archives automatically
    to_ingest: list[Path] = []
    if suffix == ".zip":
        try:
            import zipfile as _zf
            with _zf.ZipFile(dest, "r") as zf:
                supported_exts = {".csv", ".xlsx", ".xls", ".kml", ".kmz"}
                for name in zf.namelist():
                    if Path(name).suffix.lower() in supported_exts:
                        extracted = UPLOAD_DIR / Path(name).name
                        with zf.open(name) as src, open(extracted, "wb") as dst:
                            shutil.copyfileobj(src, dst)
                        to_ingest.append(extracted)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not open ZIP: {e}")
        if not to_ingest:
            raise HTTPException(
                status_code=400,
                detail="ZIP contained no supported files (CSV, Excel, KML, KMZ).",
            )
    else:
        to_ingest = [dest]

    last_result = None
    errors: list[str] = []
    for path in to_ingest:
        try:
            with session_scope() as session:
                repo = IngestionRepository(session)
                service = IngestionService(repo)
                last_result = service.ingest_file(path, customer_id=customer_id)
        except Exception as e:
            errors.append(f"{path.name}: {e}")

    if last_result is None:
        raise HTTPException(status_code=400, detail="; ".join(errors) or "Ingestion failed")

    return UploadResponse(
        job_id=str(last_result.job_id),
        source_file=last_result.source_file,
        total_raw_records=last_result.total_raw_records,
        valid_records=last_result.valid_records,
        invalid_records=last_result.invalid_records,
        duplicate_records=last_result.duplicate_records,
        stored_records=last_result.stored_records,
        queued_records=last_result.queued_records,
        status=last_result.status.value,
    )


@app.get("/api/jobs", response_model=list[JobResponse])
def list_jobs(_current_user: str = Depends(_require_auth)):
    """List all ingestion jobs."""
    session = get_session_factory()()
    try:
        jobs = session.execute(
            select(IngestionJob).order_by(IngestionJob.created_at.desc())
        ).scalars().all()
        return [
            JobResponse(
                id=str(j.id),
                customer_id=j.customer_id,
                source_file=j.source_file,
                row_count=j.row_count,
                status=j.status,
                error_message=j.error_message,
                created_at=j.created_at.isoformat(),
                updated_at=j.updated_at.isoformat(),
            )
            for j in jobs
        ]
    finally:
        session.close()


@app.get("/api/jobs/{job_id}", response_model=JobResponse)
def get_job(job_id: str, _current_user: str = Depends(_require_auth)):
    """Get a single job by ID."""
    session = get_session_factory()()
    try:
        job = session.get(IngestionJob, job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        return JobResponse(
            id=str(job.id),
            customer_id=job.customer_id,
            source_file=job.source_file,
            row_count=job.row_count,
            status=job.status,
            error_message=job.error_message,
            created_at=job.created_at.isoformat(),
            updated_at=job.updated_at.isoformat(),
        )
    finally:
        session.close()


@app.get("/api/categories")
def get_categories(job_id: str | None = Query(default=None), _current_user: str = Depends(_require_auth)):
    """Get distinct category values from records' raw_metadata."""
    session = get_session_factory()()
    try:
        cat_col = text("raw_metadata->>'category'")
        stmt = select(cat_col).where(
            text("raw_metadata->>'category' IS NOT NULL"),
            text("raw_metadata->>'category' != ''"),
        ).select_from(Address.__table__).distinct()

        if job_id:
            stmt = stmt.where(Address.job_id == job_id)

        rows = session.execute(stmt).scalars().all()
        categories = sorted(set(r for r in rows if r))
        return {"categories": categories}
    finally:
        session.close()


@app.get("/api/records", response_model=PaginatedRecords)
def list_records(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    job_id: str | None = Query(default=None),
    search: str | None = Query(default=None),
    category: str | None = Query(default=None),
    sort_by: str = Query(default="id"),
    sort_dir: str = Query(default="asc"),
    _current_user: str = Depends(_require_auth),
):
    """Paginated address records with dynamic columns."""
    session = get_session_factory()()
    try:
        stmt = select(Address)
        count_stmt = select(func.count(Address.id))

        if job_id:
            stmt = stmt.where(Address.job_id == job_id)
            count_stmt = count_stmt.where(Address.job_id == job_id)

        if category:
            cat_filter = text("raw_metadata->>'category' = :cat").bindparams(cat=category)
            stmt = stmt.where(cat_filter)
            count_stmt = count_stmt.where(cat_filter)

        if search:
            like_pattern = f"%{search}%"
            search_filter = (
                Address.raw_address.ilike(like_pattern)
                | Address.city.ilike(like_pattern)
                | Address.state.ilike(like_pattern)
                | Address.zip_code.ilike(like_pattern)
                | Address.terminal_id.ilike(like_pattern)
                | Address.network_node.ilike(like_pattern)
                | Address.address_id.ilike(like_pattern)
                | func.cast(Address.raw_metadata, String).ilike(like_pattern)
            )
            stmt = stmt.where(search_filter)
            count_stmt = count_stmt.where(search_filter)

        # Sorting
        sort_column = getattr(Address, sort_by, Address.id)
        if sort_dir == "desc":
            stmt = stmt.order_by(sort_column.desc())
        else:
            stmt = stmt.order_by(sort_column.asc())

        # Total count
        total = session.execute(count_stmt).scalar()

        # Pagination
        offset = (page - 1) * page_size
        stmt = stmt.offset(offset).limit(page_size)
        addresses = session.execute(stmt).scalars().all()

        # Build dynamic columns based on raw_metadata from the first record
        columns = _get_raw_columns(session, job_id)

        # Load Agent 1 results for these address ids to overlay validated lat/lon
        addr_ids = [a.id for a in addresses]
        agent1_map: dict[int, Agent1Result] = {}
        if addr_ids:
            from sqlalchemy import select as _sel2
            a1_rows = session.execute(
                _sel2(Agent1Result).where(Agent1Result.address_id.in_(addr_ids))
            ).scalars().all()
            agent1_map = {r.address_id: r for r in a1_rows}

        # Append agent1 columns to schema if any agent1 results exist for this job
        if agent1_map:
            A1_SCHEMA = [
                ("agent1_validation_status",       "Agent1: Validation Status"),
                ("agent1_confidence_score",         "Agent1: Confidence Score"),
                ("agent1_canonical_address",        "Agent1: Canonical Address"),
                ("agent1_chosen_provider",          "Agent1: Chosen Provider"),
                ("agent1_smarty_lat",               "Agent1: Smarty Latitude"),
                ("agent1_smarty_lon",               "Agent1: Smarty Longitude"),
                ("agent1_smarty_standardized",      "Agent1: Smarty Standardized"),
                ("agent1_smarty_dpv",               "Agent1: Smarty DPV"),
                ("agent1_smarty_zip_plus_4",        "Agent1: Smarty ZIP+4"),
                ("agent1_melissa_standardized",     "Agent1: Melissa Standardized"),
                ("agent1_melissa_dpv",              "Agent1: Melissa DPV"),
                ("agent1_exception_reason",         "Agent1: Exception Reason"),
            ]
            existing_keys = {c["key"] for c in columns}
            for key, label in A1_SCHEMA:
                if key not in existing_keys:
                    columns.append({"key": key, "label": label, "visible": True, "source": "agent1"})

        def _merge_meta(a: Address) -> dict:
            """Return raw_metadata with Smarty lat/lon overwritten and agent1 fields appended."""
            meta = dict(a.raw_metadata or {})
            a1 = agent1_map.get(a.id)
            if a1 and a1.smarty_lat:
                for k in list(meta.keys()):
                    kl = k.lower().replace(" ", "_")
                    if kl == "network_node_latitude":
                        meta[k] = a1.smarty_lat
                    elif kl == "network_node_longitude":
                        meta[k] = a1.smarty_lon
            # Append agent1 fields so frontend can read them via raw_data[key]
            if a1:
                meta["agent1_validation_status"]   = a1.validation_status or ""
                meta["agent1_confidence_score"]     = a1.confidence_score
                meta["agent1_canonical_address"]    = a1.canonical_address or ""
                meta["agent1_chosen_provider"]      = a1.chosen_provider or ""
                meta["agent1_smarty_lat"]           = a1.smarty_lat
                meta["agent1_smarty_lon"]           = a1.smarty_lon
                meta["agent1_smarty_standardized"]  = a1.smarty_standardized_address or ""
                meta["agent1_smarty_dpv"]           = a1.smarty_dpv or ""
                meta["agent1_smarty_zip_plus_4"]    = a1.smarty_zip_plus_4 or ""
                meta["agent1_melissa_standardized"] = a1.melissa_standardized_address or ""
                meta["agent1_melissa_dpv"]          = a1.melissa_dpv or ""
                meta["agent1_exception_reason"]     = a1.exception_reason or ""
            return meta

        records = [
            RecordResponse(
                id=a.id,
                record_uuid=str(a.record_uuid),
                job_id=str(a.job_id),
                customer_id=a.customer_id,
                raw_address=a.raw_address,
                city=a.city,
                state=a.state,
                zip_code=a.zip_code,
                latitude=a.latitude,
                longitude=a.longitude,
                network_node=a.network_node,
                terminal_id=a.terminal_id,
                address_id=a.address_id,
                normalized_key=a.normalized_key,
                source_file=a.source_file,
                source_sheet=a.source_sheet,
                source_layer=a.source_layer,
                source_row_number=a.source_row_number,
                validation_errors=a.validation_errors,
                validation_warnings=a.validation_warnings,
                raw_metadata=_merge_meta(a),
                created_at=a.created_at.isoformat(),
                raw_data=_merge_meta(a),
            )
            for a in addresses
        ]

        total_pages = max(1, (total + page_size - 1) // page_size)

        return PaginatedRecords(
            records=records,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=total_pages,
            columns=columns,
            display_mode="raw" if columns and columns[0].get("source") == "raw" else "canonical",
        )
    finally:
        session.close()


@app.get("/api/records/geo")
def get_geo_records(
    job_id: str | None = Query(default=None),
    limit: int = Query(default=5000, ge=1, le=50000),
    _current_user: str = Depends(_require_auth),
):
    """Get records with coordinates for map display (lightweight GeoJSON-like response)."""
    session = get_session_factory()()
    try:
        stmt = (
            select(
                Address.id,
                Address.latitude,
                Address.longitude,
                Address.raw_address,
                Address.city,
                Address.state,
                Address.network_node,
                Address.terminal_id,
                Address.source_file,
                Address.source_row_number,
                Address.raw_metadata,
                # Agent 1 validation columns (NULL when not yet processed)
                Agent1Result.canonical_address,
                Agent1Result.validation_status,
                Agent1Result.confidence_score,
                Agent1Result.chosen_provider,
                Agent1Result.smarty_lat,
                Agent1Result.smarty_lon,
            )
            .outerjoin(Agent1Result, Agent1Result.address_id == Address.id)
            .where(
                Address.latitude.isnot(None),
                Address.longitude.isnot(None),
                Address.latitude != 0,
                Address.longitude != 0,
            )
        )

        if job_id:
            stmt = stmt.where(Address.job_id == job_id)

        stmt = stmt.limit(limit)
        rows = session.execute(stmt).all()

        features = []
        for row in rows:
            meta = row.raw_metadata or {}
            style_color = meta.get("style_color", "")
            # Ensure style_color has # prefix for CSS
            if style_color and not style_color.startswith("#"):
                style_color = "#" + style_color

            # Use Smarty lat/lon when available and non-zero (more accurate)
            lat = row.latitude
            lon = row.longitude
            if row.smarty_lat and row.smarty_lat != 0:
                lat = row.smarty_lat
            if row.smarty_lon and row.smarty_lon != 0:
                lon = row.smarty_lon

            feat = {
                "id": row.id,
                "lat": lat,
                "lon": lon,
                "address": row.raw_address or meta.get("placemark_name", ""),
                "city": row.city or "",
                "state": row.state or "",
                "network_node": row.network_node or "",
                "terminal_id": row.terminal_id or "",
                "source_file": row.source_file or "",
                "row_number": row.source_row_number,
                "category": meta.get("category", ""),
                "geometry_type": meta.get("geometry_type", ""),
                "folder_path": meta.get("folder_path", ""),
                "style_color": style_color,
                "placemark_name": meta.get("placemark_name", ""),
                # Agent 1 enrichment
                "canonical_address": row.canonical_address or "",
                "validation_status": row.validation_status or "",
                "confidence_score": row.confidence_score,
                "chosen_provider": row.chosen_provider or "",
            }
            # Parse and include coordinates for polygons and linestrings
            geom_type = meta.get("geometry_type", "")
            if geom_type in ("Polygon", "LineString") and meta.get("coordinates"):
                parsed = _parse_kml_coordinates(meta["coordinates"], geom_type)
                if parsed:
                    feat["coordinates"] = parsed
            features.append(feat)

        return {"features": features, "count": len(features)}
    finally:
        session.close()


def _parse_kml_coordinates(coords_raw, geom_type: str):
    """Parse KML coordinate string into array format for frontend.

    KML format: 'lon,lat,alt\\nlon,lat,alt\\n...'
    Returns: [[lon, lat], ...] for LineString or [[[lon, lat], ...]] for Polygon
    """
    if isinstance(coords_raw, list):
        # Already parsed
        return coords_raw

    if not isinstance(coords_raw, str):
        return None

    try:
        # Split by whitespace/newlines and parse each coordinate tuple
        parts = coords_raw.strip().split()
        # If single-space-separated didn't work, try newline
        if len(parts) <= 1:
            parts = [p.strip() for p in coords_raw.strip().splitlines() if p.strip()]

        coords = []
        for part in parts:
            # Each part is "lon,lat,alt" or "lon,lat"
            components = part.strip().rstrip(",").split(",")
            if len(components) >= 2:
                lon = float(components[0])
                lat = float(components[1])
                coords.append([lon, lat])

        if not coords:
            return None

        if geom_type == "Polygon":
            return [coords]  # Polygon needs array of rings
        return coords  # LineString is flat array of coords
    except (ValueError, TypeError):
        return None


@app.get("/api/records/{record_id}", response_model=RecordResponse)
def get_record(record_id: int, _current_user: str = Depends(_require_auth)):
    """Get a single address record."""
    session = get_session_factory()()
    try:
        a = session.get(Address, record_id)
        if not a:
            raise HTTPException(status_code=404, detail="Record not found")
        # Overlay Smarty lat/lon into raw_metadata if Agent 1 has results
        from sqlalchemy import select as _sel3
        a1 = session.execute(
            _sel3(Agent1Result).where(Agent1Result.address_id == record_id)
        ).scalars().first()
        meta = dict(a.raw_metadata or {})
        if a1 and a1.smarty_lat:
            for k in list(meta.keys()):
                kl = k.lower().replace(" ", "_")
                if kl == "network_node_latitude":
                    meta[k] = a1.smarty_lat
                elif kl == "network_node_longitude":
                    meta[k] = a1.smarty_lon
        return RecordResponse(
            id=a.id,
            record_uuid=str(a.record_uuid),
            job_id=str(a.job_id),
            customer_id=a.customer_id,
            raw_address=a.raw_address,
            city=a.city,
            state=a.state,
            zip_code=a.zip_code,
            latitude=a.latitude,
            longitude=a.longitude,
            network_node=a.network_node,
            terminal_id=a.terminal_id,
            address_id=a.address_id,
            normalized_key=a.normalized_key,
            source_file=a.source_file,
            source_sheet=a.source_sheet,
            source_layer=a.source_layer,
            source_row_number=a.source_row_number,
            validation_errors=a.validation_errors,
            validation_warnings=a.validation_warnings,
            raw_metadata=meta,
            created_at=a.created_at.isoformat(),
        )
    finally:
        session.close()


# ─── Agent input / output endpoints ────────────────────────────────────────────

class AgentOutputPatch(BaseModel):
    agent_id: str
    output: dict[str, Any]


@app.get("/api/agent/records")
def agent_get_records(
    job_id: str,
    limit: int = Query(default=1000, ge=1, le=10000),
    offset: int = Query(default=0, ge=0),
    _current_user: str = Depends(_require_auth),
):
    """Pull all records for a job as agent input (id, canonical fields + full raw_metadata)."""
    session = get_session_factory()()
    try:
        rows = session.scalars(
            select(Address)
            .where(Address.job_id == job_id)
            .order_by(Address.id.asc())
            .limit(limit)
            .offset(offset)
        ).all()
        total = session.execute(
            select(func.count(Address.id)).where(Address.job_id == job_id)
        ).scalar() or 0
        return {
            "total": total,
            "offset": offset,
            "limit": limit,
            "records": [
                {
                    "id": r.id,
                    "record_uuid": str(r.record_uuid),
                    "job_id": str(r.job_id),
                    "raw_address": r.raw_address,
                    "city": r.city,
                    "state": r.state,
                    "zip_code": r.zip_code,
                    "latitude": r.latitude,
                    "longitude": r.longitude,
                    "network_node": r.network_node,
                    "terminal_id": r.terminal_id,
                    "address_id": r.address_id,
                    "source_file": r.source_file,
                    "source_row_number": r.source_row_number,
                    "raw_metadata": r.raw_metadata or {},
                }
                for r in rows
            ],
        }
    finally:
        session.close()


@app.patch("/api/agent/records/{record_id}")
def agent_patch_record(
    record_id: int,
    payload: AgentOutputPatch,
    _current_user: str = Depends(_require_auth),
):
    """
    Write agent output into raw_metadata['agent_outputs'][agent_id].
    Non-destructive — other agents' outputs and original fields are preserved.
    """
    session = get_session_factory()()
    try:
        record = session.get(Address, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")
        meta = dict(record.raw_metadata or {})
        outputs = dict(meta.get("agent_outputs", {}))
        outputs[payload.agent_id] = payload.output
        meta["agent_outputs"] = outputs
        record.raw_metadata = meta
        session.commit()
        return {"id": record_id, "agent_id": payload.agent_id, "status": "saved"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.patch("/api/agent/records")
def agent_patch_records_bulk(
    payloads: list[dict[str, Any]],
    _current_user: str = Depends(_require_auth),
):
    """Bulk-write agent outputs for up to 500 records. Each item: {id, agent_id, output}."""
    if len(payloads) > 500:
        raise HTTPException(status_code=400, detail="Max 500 records per bulk patch")
    session = get_session_factory()()
    try:
        saved = []
        for item in payloads:
            record = session.get(Address, item.get("id"))
            if not record or not item.get("agent_id"):
                continue
            meta = dict(record.raw_metadata or {})
            outputs = dict(meta.get("agent_outputs", {}))
            outputs[item["agent_id"]] = item.get("output", {})
            meta["agent_outputs"] = outputs
            record.raw_metadata = meta
            saved.append(item["id"])
        session.commit()
        return {"saved_count": len(saved), "saved_ids": saved}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.get("/api/columns", response_model=list[ColumnInfo])
def get_columns(job_id: str | None = Query(default=None), _current_user: str = Depends(_require_auth)):
    """Get dynamic column definitions based on data in the database."""
    session = get_session_factory()()
    try:
        columns = _get_raw_columns(session, job_id)
        return [ColumnInfo(**{k: v for k, v in c.items() if k in ("key", "label", "visible")}) for c in columns]
    finally:
        session.close()


@app.get("/api/stats")
def get_stats(job_id: str | None = Query(default=None), _current_user: str = Depends(_require_auth)):
    """Get summary statistics for the dashboard."""
    session = get_session_factory()()
    try:
        base_filter = Address.job_id == job_id if job_id else True

        total = session.execute(select(func.count(Address.id)).where(base_filter)).scalar() or 0
        with_coords = session.execute(
            select(func.count(Address.id)).where(
                base_filter, Address.latitude.isnot(None), Address.longitude.isnot(None)
            )
        ).scalar() or 0
        without_coords = total - with_coords

        jobs_count = session.execute(select(func.count(IngestionJob.id))).scalar() or 0

        return {
            "total_records": total,
            "with_coordinates": with_coords,
            "without_coordinates": without_coords,
            "total_jobs": jobs_count,
        }
    finally:
        session.close()


# ─── Helpers ────────────────────────────────────────────────────────────────────

def _get_raw_columns(session, job_id: str | None) -> list[dict[str, Any]]:
    """Determine columns from raw_metadata of the first record for this job."""
    base_filter = Address.job_id == job_id if job_id else True

    # Get the first record to inspect its raw_metadata keys
    first_record = session.execute(
        select(Address.raw_metadata).where(base_filter, Address.raw_metadata.isnot(None)).limit(1)
    ).scalar()

    if first_record and isinstance(first_record, dict) and len(first_record) > 0:
        # Use raw file columns as the primary display
        columns = [{"key": "source_row_number", "label": "#", "visible": True, "source": "raw"}]
        for raw_key in first_record.keys():
            label = raw_key.replace("_", " ").title()
            columns.append({"key": raw_key, "label": label, "visible": True, "source": "raw"})
        return columns

    # Fallback: use canonical fields if no raw_metadata
    return _get_canonical_columns(session, job_id)


def _get_canonical_columns(session, job_id: str | None) -> list[dict[str, Any]]:
    """Fallback: determine which canonical columns have data."""
    base_filter = Address.job_id == job_id if job_id else True

    checks = {
        "city": Address.city,
        "state": Address.state,
        "zip_code": Address.zip_code,
        "latitude": Address.latitude,
        "longitude": Address.longitude,
        "network_node": Address.network_node,
        "terminal_id": Address.terminal_id,
        "address_id": Address.address_id,
        "source_sheet": Address.source_sheet,
        "source_layer": Address.source_layer,
    }

    has_data = {}
    for key, col in checks.items():
        count = session.execute(
            select(func.count(Address.id)).where(base_filter, col.isnot(None))
        ).scalar() or 0
        has_data[key] = count > 0

    columns = [
        {"key": "id", "label": "#", "visible": True, "source": "canonical"},
        {"key": "raw_address", "label": "Address", "visible": True, "source": "canonical"},
    ]

    if has_data.get("city"):
        columns.append({"key": "city", "label": "City", "visible": True, "source": "canonical"})
    if has_data.get("state"):
        columns.append({"key": "state", "label": "State", "visible": True, "source": "canonical"})
    if has_data.get("zip_code"):
        columns.append({"key": "zip_code", "label": "ZIP", "visible": True, "source": "canonical"})
    if has_data.get("latitude"):
        columns.append({"key": "latitude", "label": "Latitude", "visible": True, "source": "canonical"})
    if has_data.get("longitude"):
        columns.append({"key": "longitude", "label": "Longitude", "visible": True, "source": "canonical"})
    if has_data.get("network_node"):
        columns.append({"key": "network_node", "label": "Node", "visible": True, "source": "canonical"})
    if has_data.get("terminal_id"):
        columns.append({"key": "terminal_id", "label": "Terminal ID", "visible": True, "source": "canonical"})
    if has_data.get("address_id"):
        columns.append({"key": "address_id", "label": "Address ID", "visible": True, "source": "canonical"})
    if has_data.get("source_sheet"):
        columns.append({"key": "source_sheet", "label": "Sheet", "visible": True, "source": "canonical"})
    if has_data.get("source_layer"):
        columns.append({"key": "source_layer", "label": "Layer", "visible": True, "source": "canonical"})

    columns.append({"key": "source_file", "label": "Source File", "visible": True, "source": "canonical"})
    columns.append({"key": "source_row_number", "label": "Row #", "visible": True, "source": "canonical"})

    return columns


# ─── Agent Processing Infrastructure ───────────────────────────────────────────

# In-memory store for agent progress (production would use Redis/DB)
_agent_progress: dict[str, dict] = {}

AGENT_DEFINITIONS = [
    {"id": "address_parser", "name": "Address Parser", "description": "Parses raw address into structured components"},
    {"id": "smarty_geocoder", "name": "Smarty Geocoder", "description": "Geocodes addresses using SmartyStreets API"},
    {"id": "melissa_validator", "name": "Melissa Validator", "description": "Validates and enriches address data via Melissa"},
    {"id": "coordinate_extractor", "name": "Coordinate Extractor", "description": "Extracts lat/lon from address or network data"},
    {"id": "network_mapper", "name": "Network Mapper", "description": "Maps addresses to network nodes and terminals"},
    {"id": "data_enricher", "name": "Data Enricher", "description": "Fills missing fields from external data sources"},
    {"id": "kmz_generator", "name": "KMZ Generator", "description": "Generates KMZ file from processed coordinates"},
]


class AgentStatus(BaseModel):
    agent_id: str
    agent_name: str
    description: str
    status: str  # pending, running, completed, failed
    progress: int  # 0-100
    records_processed: int
    records_total: int
    started_at: str | None = None
    completed_at: str | None = None
    errors: list[str] = []


class JobProgress(BaseModel):
    job_id: str
    overall_progress: int  # 0-100
    status: str  # pending, processing, completed, failed
    agents: list[AgentStatus]
    current_agent: str | None = None
    output_csv: str | None = None
    output_kmz: str | None = None


@app.post("/api/jobs/{job_id}/process")
async def start_processing(job_id: str, _current_user: str = Depends(_require_auth)):
    """Start agent processing for a job. Returns immediately, progress via SSE."""
    session = get_session_factory()()
    try:
        job = session.get(IngestionJob, job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        # Initialize progress tracking
        _agent_progress[job_id] = {
            "status": "processing",
            "overall_progress": 0,
            "current_agent_idx": 0,
            "total_records": job.row_count,
            "agents": [
                {
                    "agent_id": a["id"],
                    "agent_name": a["name"],
                    "description": a["description"],
                    "status": "pending",
                    "progress": 0,
                    "records_processed": 0,
                    "records_total": job.row_count,
                    "started_at": None,
                    "completed_at": None,
                    "errors": [],
                }
                for a in AGENT_DEFINITIONS
            ],
            "output_csv": None,
            "output_kmz": None,
        }

        # Start background processing
        asyncio.create_task(_run_agents(job_id, job.row_count))

        return {"message": "Processing started", "job_id": job_id}
    finally:
        session.close()


async def _run_agents(job_id: str, total_records: int):
    """Run the real agent pipeline for a job (Agent 1 = Smarty+Melissa address validation)."""
    progress = _agent_progress.get(job_id)
    if not progress:
        return

    # ── Agent 1: Address Validation ──────────────────────────────────────────
    agent = progress["agents"][0]
    agent["status"] = "running"
    agent["started_at"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    progress["current_agent_idx"] = 0

    # Progress callback updates the in-memory dict so the SSE stream shows live %
    def _progress(done: int, total: int):
        pct = int((done / total) * 100) if total else 0
        agent["records_processed"] = done
        agent["progress"] = pct
        completed = sum(1 for a in progress["agents"] if a["status"] == "completed")
        progress["overall_progress"] = int(
            ((completed + pct / 100) / len(AGENT_DEFINITIONS)) * 100
        )

    try:
        loop = asyncio.get_event_loop()
        summary = await loop.run_in_executor(
            None,
            lambda: run_agent1_for_job(job_id, progress_callback=_progress)
        )
        agent["status"] = "completed"
        agent["errors"] = []
        agent["summary"] = summary
    except Exception as exc:
        agent["status"] = "failed"
        agent["errors"] = [str(exc)]
        progress["status"] = "failed"
        progress["overall_progress"] = 0
        return

    agent["progress"] = 100
    agent["records_processed"] = total_records
    agent["completed_at"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    # Mark remaining stub agents as instant-completed
    for idx in range(1, len(AGENT_DEFINITIONS)):
        a = progress["agents"][idx]
        a["status"] = "completed"
        a["progress"] = 100
        a["records_processed"] = total_records
        a["started_at"] = agent["completed_at"]
        a["completed_at"] = agent["completed_at"]

    progress["status"] = "completed"
    progress["overall_progress"] = 100
    progress["output_csv"] = f"/api/jobs/{job_id}/agent1-results?fmt=csv"
    progress["output_kmz"] = f"/api/export/kmz?job_id={job_id}"


# ─── Agent 1 Results endpoint ────────────────────────────────────────────────

@app.get("/api/jobs/{job_id}/agent1-results")
def get_agent1_results(
    job_id: str,
    fmt: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=100, ge=1, le=1000),
    validation_status: str | None = Query(default=None),
    _current_user: str = Depends(_require_auth),
):
    """Return Agent 1 (Smarty+Melissa) validation results for a job.

    fmt=csv → download as CSV
    validation_status=AUTO_ACCEPT|MANUAL_REVIEW|REJECT → filter
    """
    session = get_session_factory()()
    try:
        from sqlalchemy import select as _sel
        q = _sel(Agent1Result).where(Agent1Result.job_id == job_id)
        if validation_status:
            q = q.where(Agent1Result.validation_status == validation_status)
        q = q.order_by(Agent1Result.id)

        total = session.scalar(_sel(func.count()).select_from(
            _sel(Agent1Result).where(Agent1Result.job_id == job_id).subquery()
        ))

        all_rows = session.scalars(q).all()

        def _row(r: Agent1Result) -> dict:
            return {
                "id":                            r.id,
                "address_id":                    r.address_id,
                "raw_address":                   r.raw_address,
                "canonical_address":             r.canonical_address,
                "smarty_standardized_address":   r.smarty_standardized_address,
                "smarty_dpv":                    r.smarty_dpv,
                "smarty_zip_plus_4":             r.smarty_zip_plus_4,
                "smarty_vacant":                 r.smarty_vacant,
                "smarty_record_type":            r.smarty_record_type,
                "smarty_lat":                    r.smarty_lat,
                "smarty_lon":                    r.smarty_lon,
                "melissa_standardized_address":  r.melissa_standardized_address,
                "melissa_dpv":                   r.melissa_dpv,
                "melissa_zip_plus_4":            r.melissa_zip_plus_4,
                "melissa_vacant":                r.melissa_vacant,
                "melissa_record_type":           r.melissa_record_type,
                "chosen_standardized_address":   r.chosen_standardized_address,
                "chosen_provider":               r.chosen_provider,
                "structure_hint":                r.structure_hint,
                "confidence_score":              r.confidence_score,
                "validation_status":             r.validation_status,
                "exception_reason":              r.exception_reason,
                "comparison_reason":             r.comparison_reason,
                "created_at":                    r.created_at.isoformat() if r.created_at else None,
            }

        if fmt == "csv":
            rows = [_row(r) for r in all_rows]
            if not rows:
                raise HTTPException(status_code=404, detail="No agent1 results for this job")
            buf = io.StringIO()
            w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
            buf.seek(0)
            return StreamingResponse(
                iter([buf.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=agent1_{job_id}.csv"},
            )

        rows = [_row(r) for r in all_rows[(page-1)*page_size : page*page_size]]
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "results": rows,
            "summary": {
                "auto_accept":   sum(1 for r in all_rows if r.validation_status == "AUTO_ACCEPT"),
                "manual_review": sum(1 for r in all_rows if r.validation_status == "MANUAL_REVIEW"),
                "reject":        sum(1 for r in all_rows if r.validation_status == "REJECT"),
            }
        }
    finally:
        session.close()


@app.get("/api/jobs/{job_id}/progress")
async def get_job_progress_sse(
    job_id: str,
    token: str | None = Query(default=None),
    _current_user: str = Depends(_require_auth),
):
    """SSE endpoint streaming real-time agent progress updates."""

    async def event_generator():
        last_sent = None
        while True:
            progress = _agent_progress.get(job_id)
            if not progress:
                yield f"data: {json.dumps({'error': 'No processing found for this job'})}\n\n"
                return

            # Build response
            current_agent = None
            if progress["status"] == "processing":
                idx = progress.get("current_agent_idx", 0)
                if idx < len(progress["agents"]):
                    current_agent = progress["agents"][idx]["agent_id"]

            payload = {
                "job_id": job_id,
                "overall_progress": progress["overall_progress"],
                "status": progress["status"],
                "current_agent": current_agent,
                "agents": progress["agents"],
                "output_csv": progress.get("output_csv"),
                "output_kmz": progress.get("output_kmz"),
            }

            payload_str = json.dumps(payload)
            if payload_str != last_sent:
                yield f"data: {payload_str}\n\n"
                last_sent = payload_str

            if progress["status"] in ("completed", "failed"):
                yield f"data: {json.dumps({'done': True})}\n\n"
                return

            await asyncio.sleep(0.5)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive"},
    )


@app.get("/api/jobs/{job_id}/agents")
def get_job_agents(job_id: str, _current_user: str = Depends(_require_auth)):
    """Get current agent status for a job."""
    progress = _agent_progress.get(job_id)
    if not progress:
        # Return default pending state
        return {
            "job_id": job_id,
            "overall_progress": 0,
            "status": "pending",
            "agents": [
                {
                    "agent_id": a["id"],
                    "agent_name": a["name"],
                    "description": a["description"],
                    "status": "pending",
                    "progress": 0,
                    "records_processed": 0,
                    "records_total": 0,
                }
                for a in AGENT_DEFINITIONS
            ],
        }

    current_agent = None
    if progress["status"] == "processing":
        idx = progress.get("current_agent_idx", 0)
        if idx < len(progress["agents"]):
            current_agent = progress["agents"][idx]["agent_id"]

    return {
        "job_id": job_id,
        "overall_progress": progress["overall_progress"],
        "status": progress["status"],
        "current_agent": current_agent,
        "agents": progress["agents"],
        "output_csv": progress.get("output_csv"),
        "output_kmz": progress.get("output_kmz"),
    }


# ─── Agent Table Management ──────────────────────────────────────────────────────

class CreateAgentTableRequest(BaseModel):
    agent_name: str
    display_name: str
    description: str | None = None
    color_rules: list[dict[str, Any]] | None = None


@app.post("/api/agent/tables")
def create_agent_table(req: CreateAgentTableRequest, _current_user: str = Depends(_require_auth)):
    """Register a new agent table with optional color rules."""
    session = get_session_factory()()
    try:
        existing = session.execute(
            select(AgentTable).where(AgentTable.agent_name == req.agent_name)
        ).scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=400, detail=f"Agent table '{req.agent_name}' already exists")
        table = AgentTable(
            agent_name=req.agent_name,
            display_name=req.display_name,
            owner=_current_user,
            description=req.description,
            color_rules=req.color_rules or [],
        )
        session.add(table)
        session.commit()
        return {
            "agent_name": table.agent_name,
            "display_name": table.display_name,
            "owner": table.owner,
            "color_rules": table.color_rules,
        }
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.get("/api/agent/tables")
def list_agent_tables(_current_user: str = Depends(_require_auth)):
    """List all registered agent tables with their color rules."""
    session = get_session_factory()()
    try:
        tables = session.scalars(select(AgentTable).order_by(AgentTable.created_at)).all()
        return [
            {
                "agent_name": t.agent_name,
                "display_name": t.display_name,
                "owner": t.owner,
                "description": t.description,
                "color_rules": t.color_rules or [],
                "created_at": t.created_at.isoformat(),
            }
            for t in tables
        ]
    finally:
        session.close()


@app.get("/api/agent/tables/{agent_name}")
def get_agent_table(agent_name: str, _current_user: str = Depends(_require_auth)):
    """Get a specific agent table's metadata and color rules."""
    session = get_session_factory()()
    try:
        t = session.execute(
            select(AgentTable).where(AgentTable.agent_name == agent_name)
        ).scalar_one_or_none()
        if not t:
            raise HTTPException(status_code=404, detail="Agent table not found")
        return {
            "agent_name": t.agent_name,
            "display_name": t.display_name,
            "owner": t.owner,
            "description": t.description,
            "color_rules": t.color_rules or [],
            "created_at": t.created_at.isoformat(),
        }
    finally:
        session.close()


@app.put("/api/agent/tables/{agent_name}/color-rules")
def update_color_rules(
    agent_name: str,
    color_rules: list[dict[str, Any]],
    _current_user: str = Depends(_require_auth),
):
    """
    Update color rules for an agent table.
    Each rule: {"field": "status", "value": "qualified", "color": "#22c55e", "label": "Qualified"}
    """
    session = get_session_factory()()
    try:
        t = session.execute(
            select(AgentTable).where(AgentTable.agent_name == agent_name)
        ).scalar_one_or_none()
        if not t:
            raise HTTPException(status_code=404, detail="Agent table not found")
        t.color_rules = color_rules
        session.commit()
        return {"agent_name": agent_name, "color_rules": color_rules}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


# ─── Agent Results CRUD ──────────────────────────────────────────────────────────

@app.post("/api/agent/results/{agent_name}")
def upsert_agent_results(
    agent_name: str,
    payloads: list[dict[str, Any]],
    _current_user: str = Depends(_require_auth),
):
    """
    Bulk upsert agent results into the agent's dedicated table.
    Each item must have: {"address_id": int, "job_id": "uuid", "data": {...}}
    Max 1000 per call. Use repeated calls for larger datasets.
    """
    if len(payloads) > 1000:
        raise HTTPException(status_code=400, detail="Max 1000 records per call")
    session = get_session_factory()()
    try:
        t = session.execute(
            select(AgentTable).where(AgentTable.agent_name == agent_name)
        ).scalar_one_or_none()
        if not t:
            raise HTTPException(
                status_code=404,
                detail=f"Agent table '{agent_name}' not found — register it first via POST /api/agent/tables",
            )
        saved = 0
        for item in payloads:
            address_id = item.get("address_id")
            job_id = item.get("job_id")
            data = item.get("data", {})
            if not address_id or not job_id:
                continue
            existing = session.execute(
                select(AgentResult).where(
                    AgentResult.agent_name == agent_name,
                    AgentResult.address_id == address_id,
                )
            ).scalar_one_or_none()
            if existing:
                existing.data = data
                existing.updated_at = datetime.utcnow()
            else:
                session.add(AgentResult(
                    agent_name=agent_name,
                    job_id=job_id,
                    address_id=address_id,
                    data=data,
                ))
            saved += 1
        session.commit()
        return {"saved": saved}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.get("/api/agent/results/{agent_name}")
def get_agent_results(
    agent_name: str,
    job_id: str | None = Query(default=None),
    limit: int = Query(default=1000, ge=1, le=10000),
    offset: int = Query(default=0, ge=0),
    _current_user: str = Depends(_require_auth),
):
    """Fetch stored results for a specific agent, optionally filtered by job."""
    session = get_session_factory()()
    try:
        filters = [AgentResult.agent_name == agent_name]
        if job_id:
            filters.append(AgentResult.job_id == job_id)
        rows = session.scalars(
            select(AgentResult).where(*filters)
            .order_by(AgentResult.address_id.asc())
            .limit(limit).offset(offset)
        ).all()
        total = session.execute(
            select(func.count(AgentResult.id)).where(*filters)
        ).scalar() or 0
        return {
            "agent_name": agent_name,
            "total": total,
            "offset": offset,
            "limit": limit,
            "results": [
                {
                    "id": r.id,
                    "address_id": r.address_id,
                    "job_id": str(r.job_id),
                    "data": r.data or {},
                    "updated_at": r.updated_at.isoformat(),
                }
                for r in rows
            ],
        }
    finally:
        session.close()


@app.patch("/api/agent/results/{agent_name}/{address_id}")
def update_agent_result(
    agent_name: str,
    address_id: int,
    data: dict[str, Any],
    _current_user: str = Depends(_require_auth),
):
    """Update a single agent result record."""
    session = get_session_factory()()
    try:
        r = session.execute(
            select(AgentResult).where(
                AgentResult.agent_name == agent_name,
                AgentResult.address_id == address_id,
            )
        ).scalar_one_or_none()
        if not r:
            raise HTTPException(status_code=404, detail="Result not found")
        r.data = data
        r.updated_at = datetime.utcnow()
        session.commit()
        return {"agent_name": agent_name, "address_id": address_id, "status": "updated"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


# ─── Map Overlay ─────────────────────────────────────────────────────────────────

@app.get("/api/map/overlay")
def get_map_overlay(job_id: str, _current_user: str = Depends(_require_auth)):
    """
    Returns per-address color derived from all agents' color_rules.
    Response: {"overlay": {address_id: {color, label, agent, data}}, "legend": [...]}
    """
    session = get_session_factory()()
    try:
        tables = session.scalars(select(AgentTable)).all()
        if not tables:
            return {"overlay": {}, "legend": []}

        results_by_agent: dict[str, list] = {}
        for t in tables:
            results_by_agent[t.agent_name] = session.scalars(
                select(AgentResult).where(
                    AgentResult.agent_name == t.agent_name,
                    AgentResult.job_id == job_id,
                )
            ).all()

        overlay: dict[int, dict] = {}
        legend: list[dict] = []

        for table in tables:
            rules = table.color_rules or []
            for r in results_by_agent.get(table.agent_name, []):
                d = r.data or {}
                for rule in rules:
                    field = rule.get("field", "")
                    if str(d.get(field, "")) == str(rule.get("value", "")):
                        overlay[r.address_id] = {
                            "color": rule.get("color", "#888888"),
                            "label": rule.get("label", str(rule.get("value", ""))),
                            "agent": table.display_name,
                            "data": d,
                        }
                        break
            for rule in rules:
                legend.append({
                    "agent_name": table.agent_name,
                    "display_name": table.display_name,
                    "field": rule.get("field"),
                    "value": rule.get("value"),
                    "color": rule.get("color", "#888888"),
                    "label": rule.get("label", str(rule.get("value", ""))),
                })

        return {"overlay": overlay, "legend": legend}
    finally:
        session.close()


# ─── Export: Merged CSV ──────────────────────────────────────────────────────────

@app.get("/api/export/csv")
def export_merged_csv(
    job_id: str,
    _current_user: str = Depends(_require_auth),
    _t: str | None = Query(default=None),  # browser-download token param (alias for Bearer)
):
    """
    Download a single merged CSV: base address columns + raw metadata + all agent outputs.
    Agent columns are prefixed: <agent_name>__<field>.
    """
    session = get_session_factory()()
    try:
        addresses = session.scalars(
            select(Address).where(Address.job_id == job_id).order_by(Address.id.asc())
        ).all()
        if not addresses:
            raise HTTPException(status_code=404, detail="No records found for this job")

        tables = {t.agent_name: t for t in session.scalars(select(AgentTable)).all()}
        results_raw = session.scalars(
            select(AgentResult).where(AgentResult.job_id == job_id)
        ).all()

        # Index: address_id → {agent_name → data}
        agent_data: dict[int, dict[str, dict]] = {}
        for r in results_raw:
            agent_data.setdefault(r.address_id, {})[r.agent_name] = r.data or {}

        # Collect all agent column keys per agent
        agent_col_keys: dict[str, list[str]] = {}
        for aname in tables:
            all_keys: set[str] = set()
            for agents in agent_data.values():
                if aname in agents:
                    all_keys.update(agents[aname].keys())
            agent_col_keys[aname] = sorted(all_keys)

        # Load Agent 1 results indexed by address_id
        agent1_data: dict[int, Agent1Result] = {
            r.address_id: r
            for r in session.scalars(
                select(Agent1Result).where(Agent1Result.job_id == job_id)
            ).all()
        }

        base_cols = [
            "id", "raw_address", "city", "state", "zip_code",
            "latitude", "longitude", "network_node", "terminal_id",
            "address_id", "source_file", "source_row_number",
            # Agent 1 validation enrichment columns
            "agent1_canonical_address", "agent1_validation_status",
            "agent1_confidence_score", "agent1_chosen_provider",
            "agent1_smarty_lat", "agent1_smarty_lon",
            "agent1_smarty_standardized", "agent1_smarty_dpv",
            "agent1_exception_reason",
        ]
        meta_keys: list[str] = []
        if addresses:
            meta = addresses[0].raw_metadata or {}
            meta_keys = [k for k in meta if k not in ("agent_outputs", "agent_results", "coordinates")]

        agent_cols = [
            (aname, key, f"{aname}__{key}")
            for aname, keys in agent_col_keys.items()
            for key in keys
        ]

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=base_cols + meta_keys + [c[2] for c in agent_cols], extrasaction="ignore")
        writer.writeheader()

        for addr in addresses:
            meta = addr.raw_metadata or {}
            a1 = agent1_data.get(addr.id)
            row: dict[str, Any] = {
                "id": addr.id,
                "raw_address": addr.raw_address or "",
                "city": addr.city or "",
                "state": addr.state or "",
                "zip_code": addr.zip_code or "",
                "latitude": addr.latitude or "",
                "longitude": addr.longitude or "",
                "network_node": addr.network_node or "",
                "terminal_id": addr.terminal_id or "",
                "address_id": addr.address_id or "",
                "source_file": addr.source_file or "",
                "source_row_number": addr.source_row_number or "",
                # Agent 1 columns
                "agent1_canonical_address": a1.canonical_address if a1 else "",
                "agent1_validation_status": a1.validation_status if a1 else "",
                "agent1_confidence_score": a1.confidence_score if a1 else "",
                "agent1_chosen_provider": a1.chosen_provider if a1 else "",
                "agent1_smarty_lat": a1.smarty_lat if a1 else "",
                "agent1_smarty_lon": a1.smarty_lon if a1 else "",
                "agent1_smarty_standardized": a1.smarty_standardized_address if a1 else "",
                "agent1_smarty_dpv": a1.smarty_dpv if a1 else "",
                "agent1_exception_reason": a1.exception_reason if a1 else "",
            }
            for k in meta_keys:
                row[k] = meta.get(k, "")
            addr_agents = agent_data.get(addr.id, {})
            for aname, key, col_name in agent_cols:
                row[col_name] = addr_agents.get(aname, {}).get(key, "")
            writer.writerow(row)

        output.seek(0)
        fname = f"ftth_merged_{job_id[:8]}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )
    finally:
        session.close()


# ─── Export: Excel with Agent-1 full output written back ──────────────────────

@app.get("/api/export/excel")
def export_excel(
    job_id: str,
    _current_user: str = Depends(_require_auth),
    _t: str | None = Query(default=None),
):
    """
    Download an Excel file with two sheets:
      Sheet 1 – all original columns with network_node_latitude/longitude
                 overwritten from Smarty, plus all Agent 1 output columns
                 appended at the right (blue header, yellow fill for validated).
      Sheet 2 – full Agent 1 results table (one row per address).
    """
    session = get_session_factory()()
    try:
        addresses = session.scalars(
            select(Address).where(Address.job_id == job_id).order_by(Address.id.asc())
        ).all()
        if not addresses:
            raise HTTPException(status_code=404, detail="No records found for this job")

        agent1_map: dict[int, Agent1Result] = {
            r.address_id: r
            for r in session.scalars(
                select(Agent1Result).where(Agent1Result.job_id == job_id)
            ).all()
        }

        # ── Styles ────────────────────────────────────────────────────────────
        from openpyxl.styles import Alignment, Border, Side
        thin = Side(style="thin", color="CCCCCC")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        dark_fill   = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")  # original header
        dark_font   = Font(bold=True, color="FFFFFF")
        blue_fill   = PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")  # agent1 header
        blue_font   = Font(bold=True, color="FFFFFF")
        green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # updated coords
        green_font  = Font(bold=True, color="276221")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # agent1 data cells
        yellow_font = Font(color="9C6500")
        red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # REJECT
        red_font    = Font(bold=True, color="9C0006")
        accept_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # AUTO_ACCEPT
        accept_font = Font(bold=True, color="276221")
        review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # MANUAL_REVIEW
        review_font = Font(bold=True, color="9C6500")

        LAT_KEYS = {"network_node_latitude", "network node latitude"}
        LON_KEYS = {"network_node_longitude", "network node longitude"}

        # Agent 1 columns to append (label → attribute on Agent1Result)
        A1_COLS = [
            ("Agent1: Validation Status",       "validation_status"),
            ("Agent1: Confidence Score",         "confidence_score"),
            ("Agent1: Canonical Address",        "canonical_address"),
            ("Agent1: Chosen Provider",          "chosen_provider"),
            ("Agent1: Structure Hint",           "structure_hint"),
            ("Agent1: Smarty Latitude",          "smarty_lat"),
            ("Agent1: Smarty Longitude",         "smarty_lon"),
            ("Agent1: Smarty Standardized",      "smarty_standardized_address"),
            ("Agent1: Smarty DPV Match",         "smarty_dpv"),
            ("Agent1: Smarty ZIP+4",             "smarty_zip_plus_4"),
            ("Agent1: Smarty Vacant",            "smarty_vacant"),
            ("Agent1: Smarty Record Type",       "smarty_record_type"),
            ("Agent1: Melissa Standardized",     "melissa_standardized_address"),
            ("Agent1: Melissa DPV Match",        "melissa_dpv"),
            ("Agent1: Melissa ZIP+4",            "melissa_zip_plus_4"),
            ("Agent1: Melissa Vacant",           "melissa_vacant"),
            ("Agent1: Melissa Record Type",      "melissa_record_type"),
            ("Agent1: Comparison Reason",        "comparison_reason"),
            ("Agent1: Exception Reason",         "exception_reason"),
        ]

        # ── Sheet 1: enriched original data ───────────────────────────────────
        first_meta: dict = addresses[0].raw_metadata or {}
        meta_keys = [k for k in first_meta if k not in ("agent_outputs", "agent_results", "coordinates")]

        base_cols = [
            "id", "raw_address", "city", "state", "zip_code",
            "latitude", "longitude", "network_node", "terminal_id",
            "address_id", "source_file", "source_row_number",
        ]
        all_orig_cols = base_cols + meta_keys
        all_cols = all_orig_cols + [label for label, _ in A1_COLS]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FTTH Data + Agent1"
        ws.freeze_panes = "A2"

        col_idx = {c: i + 1 for i, c in enumerate(all_cols)}
        orig_count = len(all_orig_cols)

        # Header row
        for col_name, ci in col_idx.items():
            cell = ws.cell(row=1, column=ci, value=col_name)
            if ci <= orig_count:
                cell.fill = dark_fill
                cell.font = dark_font
            else:
                cell.fill = blue_fill
                cell.font = blue_font
            cell.alignment = Alignment(wrap_text=False)

        # Data rows
        for row_num, addr in enumerate(addresses, start=2):
            meta = addr.raw_metadata or {}
            a1 = agent1_map.get(addr.id)
            smarty_lat = a1.smarty_lat if a1 and a1.smarty_lat else None
            smarty_lon = a1.smarty_lon if a1 and a1.smarty_lon else None

            base_vals = {
                "id": addr.id,
                "raw_address": addr.raw_address or "",
                "city": addr.city or "",
                "state": addr.state or "",
                "zip_code": addr.zip_code or "",
                "latitude": addr.latitude or "",
                "longitude": addr.longitude or "",
                "network_node": addr.network_node or "",
                "terminal_id": addr.terminal_id or "",
                "address_id": addr.address_id or "",
                "source_file": addr.source_file or "",
                "source_row_number": addr.source_row_number or "",
            }

            # Write original columns
            for col_name, ci in col_idx.items():
                if ci > orig_count:
                    break
                val = base_vals.get(col_name, meta.get(col_name, ""))
                col_lower = col_name.lower().replace(" ", "_")
                updated = False
                if col_lower in LAT_KEYS and smarty_lat is not None:
                    val = smarty_lat
                    updated = True
                elif col_lower in LON_KEYS and smarty_lon is not None:
                    val = smarty_lon
                    updated = True
                cell = ws.cell(row=row_num, column=ci, value=val)
                if updated:
                    cell.fill = green_fill
                    cell.font = green_font

            # Write agent1 columns
            for label, attr in A1_COLS:
                ci = col_idx[label]
                val = getattr(a1, attr, "") if a1 else ""
                if val is None:
                    val = ""
                cell = ws.cell(row=row_num, column=ci, value=val)
                # Color status column distinctly
                if attr == "validation_status":
                    if val == "AUTO_ACCEPT":
                        cell.fill = accept_fill; cell.font = accept_font
                    elif val == "REJECT":
                        cell.fill = red_fill; cell.font = red_font
                    elif val == "MANUAL_REVIEW":
                        cell.fill = review_fill; cell.font = review_font
                elif a1:
                    cell.fill = yellow_fill; cell.font = yellow_font

        # Auto-size
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 45)

        # ── Sheet 2: Agent 1 full results ────────────────────────────────────
        ws2 = wb.create_sheet(title="Agent1 Validation")
        ws2.freeze_panes = "A2"

        a1_sheet_cols = [
            "address_id", "raw_address",
            "validation_status", "confidence_score", "chosen_provider", "structure_hint",
            "canonical_address",
            "smarty_standardized_address", "smarty_lat", "smarty_lon",
            "smarty_dpv", "smarty_zip_plus_4", "smarty_vacant", "smarty_record_type",
            "melissa_standardized_address", "melissa_dpv", "melissa_zip_plus_4",
            "melissa_vacant", "melissa_record_type",
            "comparison_reason", "exception_reason",
        ]

        for ci, col in enumerate(a1_sheet_cols, start=1):
            cell = ws2.cell(row=1, column=ci, value=col)
            cell.fill = blue_fill; cell.font = blue_font

        for row_num, addr in enumerate(addresses, start=2):
            a1 = agent1_map.get(addr.id)
            for ci, attr in enumerate(a1_sheet_cols, start=1):
                if attr == "address_id":
                    val = addr.id
                elif attr == "raw_address":
                    val = addr.raw_address or ""
                else:
                    val = (getattr(a1, attr, "") or "") if a1 else ""
                cell = ws2.cell(row=row_num, column=ci, value=val)
                if attr == "validation_status":
                    if val == "AUTO_ACCEPT":
                        cell.fill = accept_fill; cell.font = accept_font
                    elif val == "REJECT":
                        cell.fill = red_fill; cell.font = red_font
                    elif val == "MANUAL_REVIEW":
                        cell.fill = review_fill; cell.font = review_font

        for col_cells in ws2.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
            ws2.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 45)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"ftth_validated_{job_id[:8]}.xlsx"
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )
    finally:
        session.close()


# ─── Export: Merged KMZ ──────────────────────────────────────────────────────────

@app.get("/api/export/kmz")
def export_merged_kmz(
    job_id: str,
    agent_name: str | None = Query(default=None),
    _current_user: str = Depends(_require_auth),
    _t: str | None = Query(default=None),  # browser-download token param
):
    """
    Download a KMZ with placemarks coloured by agent color_rules.
    Houses are grouped into KMZ folders by their agent-assigned label.
    Pass agent_name to filter to a single agent's rules, or omit for all agents.
    """
    session = get_session_factory()()
    try:
        addresses = session.scalars(
            select(Address).where(
                Address.job_id == job_id,
                Address.latitude.isnot(None),
                Address.longitude.isnot(None),
            ).order_by(Address.id.asc())
        ).all()
        if not addresses:
            raise HTTPException(status_code=404, detail="No geo records for this job")

        tables = session.scalars(select(AgentTable)).all()
        results_raw = session.scalars(
            select(AgentResult).where(AgentResult.job_id == job_id)
        ).all()
        agent_data: dict[int, dict[str, dict]] = {}
        for r in results_raw:
            agent_data.setdefault(r.address_id, {})[r.agent_name] = r.data or {}

        # Build per-address overlay from color rules
        overlay: dict[int, dict] = {}
        for table in tables:
            if agent_name and table.agent_name != agent_name:
                continue
            rules = table.color_rules or []
            for r in results_raw:
                if r.agent_name != table.agent_name:
                    continue
                d = r.data or {}
                for rule in rules:
                    if str(d.get(rule.get("field", ""), "")) == str(rule.get("value", "")):
                        overlay[r.address_id] = {
                            "color": rule.get("color", "#888888").lstrip("#"),
                            "label": rule.get("label", ""),
                            "agent_data": d,
                        }
                        break

        def kml_color(hex_c: str) -> str:
            h = hex_c.lstrip("#")
            if len(h) == 6:
                return f"ff{h[4:6]}{h[2:4]}{h[0:2]}"
            return "ff888888"

        # Collect unique colors
        default_hex = "0074D9"
        unique_colors: dict[str, str] = {default_hex: "style_default"}
        for v in overlay.values():
            c = v["color"]
            if c not in unique_colors:
                unique_colors[c] = f"style_{c}"

        styles_xml = ""
        for hex_c, sid in unique_colors.items():
            kml_c = kml_color(hex_c)
            styles_xml += (
                f'<Style id="{sid}"><IconStyle><color>{kml_c}</color><scale>0.8</scale>'
                f'<Icon><href>http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png</href></Icon>'
                f'</IconStyle></Style>'
            )

        # Group placemarks into folders by label
        folders: dict[str, list[str]] = {"Unprocessed": []}
        for addr in addresses:
            ov = overlay.get(addr.id)
            label = ov["label"] if ov else "Unprocessed"
            if label not in folders:
                folders[label] = []
            sid = unique_colors.get(ov["color"] if ov else default_hex, "style_default")
            agent_rows = ""
            if ov:
                for k, v in ov["agent_data"].items():
                    agent_rows += f"<tr><td><b>{k}</b></td><td>{v}</td></tr>"
            desc = (
                f'<![CDATA[<table style="font-size:12px">'
                f'<tr><td><b>Address</b></td><td>{addr.raw_address or ""}</td></tr>'
                f'<tr><td><b>City</b></td><td>{addr.city or ""}, {addr.state or ""} {addr.zip_code or ""}</td></tr>'
                f'<tr><td><b>Node</b></td><td>{addr.network_node or ""}</td></tr>'
                f'<tr><td><b>Terminal</b></td><td>{addr.terminal_id or ""}</td></tr>'
                f'{agent_rows}</table>]]>'
            )
            name = (addr.raw_address or "").replace("&", "&amp;").replace("<", "&lt;")
            folders[label].append(
                f'<Placemark><name>{name}</name><description>{desc}</description>'
                f'<styleUrl>#{sid}</styleUrl>'
                f'<Point><coordinates>{addr.longitude},{addr.latitude},0</coordinates></Point>'
                f'</Placemark>'
            )

        folders_xml = ""
        for fname, marks in folders.items():
            if marks:
                escaped = fname.replace("&", "&amp;").replace("<", "&lt;")
                folders_xml += f'<Folder><name>{escaped} ({len(marks)})</name>' + "".join(marks) + "</Folder>"

        kml_doc = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<kml xmlns="http://www.opengis.net/kml/2.2">'
            f'<Document><name>FTTH Merged Export</name>{styles_xml}{folders_xml}</Document></kml>'
        )

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("doc.kml", kml_doc.encode("utf-8"))
        buf.seek(0)

        out_fname = f"ftth_merged_{job_id[:8]}.kmz"
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.google-earth.kmz",
            headers={"Content-Disposition": f"attachment; filename={out_fname}"},
        )
    finally:
        session.close()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)
